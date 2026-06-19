"""
Excel Builder
Generates a multi-sheet Excel workbook from collected inventory data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Header style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(fgColor="2C3E50", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Sheet order (matches RVTools convention)
SHEET_ORDER = [
    "vSummary", "vInfo", "vCPU", "vMemory", "vDisk", "vNetwork",
    "vHost", "vCluster", "vDatastore", "vSwitch",
]

# Leading characters that spreadsheet apps may interpret as a formula.
# Inventory values come from FusionCompute (a VM name/description could be
# attacker-controlled), so neutralize them to prevent CSV/Excel injection.
_FORMULA_TRIGGERS = ("=", "+", "-", "@")


def _sanitize_cell(value):
    """Defang spreadsheet-formula injection in string cell values.

    A string beginning with '=', '+', '-' or '@' is prefixed with a single
    quote so Excel/Sheets render it as literal text instead of evaluating it.
    Non-string values are returned unchanged.
    """
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def build_excel(data, output_path):
    """
    Build a multi-sheet Excel workbook.

    Args:
        data: dict from InventoryCollector.collect_all()
              Keys are sheet names, values are lists of dicts.
        output_path: file path to write the .xlsx file.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in SHEET_ORDER:
        rows = data.get(sheet_name, [])
        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            ws.append(["No data collected"])
            continue

        # Write header row - union of all keys across all rows (preserve order)
        seen = set()
        headers = []
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    headers.append(key)
        ws.append(headers)

        # Write data rows (sanitized against formula injection)
        for row in rows:
            ws.append([_sanitize_cell(row.get(h, "")) for h in headers])

        # Apply header styling
        _style_header(ws)

        # Auto-fit column widths
        _auto_size_columns(ws, len(rows))

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Freeze top row
        ws.freeze_panes = "A2"

    wb.save(output_path)


def _style_header(ws):
    """Apply styling to the header row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_size_columns(ws, row_count):
    """Auto-size columns based on content width (sample first 100 rows)."""
    sample_rows = min(row_count + 1, 100)

    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=sample_rows), 1):
        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length

        # Add padding and cap width
        adjusted = min(max_length + 3, 50)
        adjusted = max(adjusted, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted
