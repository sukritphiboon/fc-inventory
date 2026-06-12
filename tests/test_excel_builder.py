"""Tests for the Excel workbook builder."""

from openpyxl import load_workbook

from excel_builder import build_excel, SHEET_ORDER


def test_build_excel_creates_all_sheets(tmp_path):
    data = {
        "vInfo": [
            {"VM Name": "a", "CPUs": 2},
            {"VM Name": "b", "CPUs": 4, "Extra": "x"},
        ]
    }
    out = tmp_path / "out.xlsx"
    build_excel(data, str(out))
    assert out.exists()

    wb = load_workbook(out)
    # Every sheet in the fixed order must exist, even when empty.
    assert wb.sheetnames == SHEET_ORDER


def test_build_excel_header_is_union_of_keys(tmp_path):
    data = {
        "vInfo": [
            {"VM Name": "a", "CPUs": 2},
            {"VM Name": "b", "CPUs": 4, "Extra": "x"},
        ]
    }
    out = tmp_path / "out.xlsx"
    build_excel(data, str(out))

    ws = load_workbook(out)["vInfo"]
    headers = [c.value for c in ws[1]]
    assert headers == ["VM Name", "CPUs", "Extra"]
    # Second data row carries the extra column; first row leaves it blank.
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("a", 2, None) or rows[0] == ("a", 2, "")
    assert rows[1][2] == "x"


def test_build_excel_empty_sheet_has_placeholder(tmp_path):
    out = tmp_path / "out.xlsx"
    build_excel({}, str(out))
    ws = load_workbook(out)["vInfo"]
    assert ws["A1"].value == "No data collected"
