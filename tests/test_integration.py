"""End-to-end pipeline test: mocked collection -> Excel -> reload and verify.

This guards the seams between the collector and the workbook builder that the
focused unit tests don't exercise together.
"""

from openpyxl import load_workbook

from collector import InventoryCollector
from excel_builder import build_excel, SHEET_ORDER
from tests.test_collector_builders import FakeClient


def test_collect_to_excel_roundtrip(tmp_path):
    collector = InventoryCollector("host", "user", "pass")
    collector.client = FakeClient()

    data = collector.collect_all()

    out = tmp_path / "inventory.xlsx"
    build_excel(data, str(out))
    assert out.exists()

    wb = load_workbook(out)
    assert wb.sheetnames == SHEET_ORDER

    # vInfo sheet should carry the single collected VM.
    vinfo = wb["vInfo"]
    headers = [c.value for c in vinfo[1]]
    assert "VM Name" in headers
    name_col = headers.index("VM Name")
    first_data_row = list(vinfo.iter_rows(min_row=2, values_only=True))[0]
    assert first_data_row[name_col] == "vm-a"

    # vDatastore should reflect the computed Used %.
    vds = wb["vDatastore"]
    ds_headers = [c.value for c in vds[1]]
    used_col = ds_headers.index("Used %")
    ds_row = list(vds.iter_rows(min_row=2, values_only=True))[0]
    assert ds_row[used_col] == 50.0
